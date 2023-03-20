import time
from functools import wraps

from django.utils.text import slugify
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from erp_demo.dox_mng.models import Document, DocumentEditPurgatory
from erp_demo.hr_mng.models import Employee, Positions, AccessLevels, \
    AccessRights, PositionsToAccessLevels, Trainings, EmployeeToTrainings
from erp_demo.main_app import custom_collections
from erp_demo.main_app.models import CaptainsLog, Requirements
from erp_demo.main_app.translator import translate_to_maimunica
from erp_demo.process_mng.models import ProcessStepToDocuments, \
    ProcessStep, Process


# For forms that select specific properties before showing the results
# -----------------------------------------------------------------------

class SupportFunctions:
    @staticmethod
    def extract_entry_by_choice(table, column_name, choice):
        if choice == 'All':
            extracted_documents = table.objects.all()
            return extracted_documents
        data = {column_name: choice}
        extracted_documents = table.objects.filter(**data)

        return extracted_documents

# Delete whole db (the order is chosen in order not have issues with the links between tables)
# -----------------------------------------------------------------------

    @staticmethod
    def delete_database():
        # M2M tables
        ProcessStepToDocuments.objects.all().delete()
        PositionsToAccessLevels.objects.all().delete()
        EmployeeToTrainings.objects.all().delete()

        # tables with no dependencies to other tables
        AccessLevels.objects.all().delete()
        AccessRights.objects.all().delete()
        Trainings.objects.all().delete()

        # tables with dependencies from other tables
        Positions.objects.all().delete()
        Employee.objects.all().delete()
        ProcessStep.objects.all().delete()
        Document.objects.all().delete()

        Process.objects.all().delete()

        CaptainsLog.objects.all().delete()

        Requirements.objects.all().delete()

        return 'Successfully deleted'

# Sort process steps
# -----------------------------------------------------------------------

    @staticmethod
    def sort_process_steps(process_obj, process_step_obj, choice):
        p_list = []

        if choice is None:
            p_list = []

        elif choice == 'All':
            for process in range(len(process_obj.objects.all())):
                p_list.append([])

                for process_step in process_step_obj.objects.all():
                    if process_step.parent_process.id == process_obj.objects.all()[process].id:
                        p_list[process].append(process_step)

        elif choice != 'All':
            chosen_process = process_obj.objects.filter(number=choice).get()
            p_list.append([])

            print(chosen_process)

            for process_step in process_step_obj.objects.all():
                if process_step.parent_process.number == chosen_process.number:
                    p_list[0].append(process_step)

        return p_list

# Get list of process steps for a process
# -----------------------------------------------------------------------

    @staticmethod
    def get_process_step_list(process, process_step_obj):
        process_step_list = []

        for process_step in process_step_obj.objects.all():
            if process_step.parent_process.id == process.id:
                process_step_list.append(process_step)

        return process_step_list

# Dict with all employees and their owned processes in a list
# -----------------------------------------------------------------------

    @staticmethod
    def get_owned_processes_list(employees, processes):
        owned_processes_dict = {
            employee: [] for employee in employees.objects.all()
        }

        for process in processes.objects.all():
            if process.process_owner in owned_processes_dict.keys():
                owned_processes_dict[process.process_owner].append(process)

        print(owned_processes_dict)
        return owned_processes_dict

# Upload to db
# -----------------------------------------------------------------------

    @staticmethod
    def recreate_database(request_object):
        # Select the file
        current_file_path = request_object['select_file']
        workbook = load_workbook(current_file_path)
        worksheet = workbook[workbook.sheetnames[0]]

        # Get the collections
        info_to_update = {}
        list_of_keys = custom_collections.list_of_keys
        coordinates = custom_collections.coordinates

        # Fill the coordinates
        consecutive_empty_rows = 0
        current_row = 1
        current_table = ''
        is_table_header_done = False
        while consecutive_empty_rows < 2:
            row_value = worksheet['A' + str(current_row)].value

            if row_value is None:
                consecutive_empty_rows += 1
                current_row += 1
                is_table_header_done = False
                current_table = ''
                continue

            if current_table == '':
                if row_value in coordinates.keys():
                    current_table = row_value
                    consecutive_empty_rows = 0
                    current_row += 1
                    continue

            if not is_table_header_done:
                current_col = 1
                while 1:
                    col_value = worksheet[get_column_letter(current_col) + str(current_row)].value
                    if col_value is None:
                        is_table_header_done = True
                        current_col -= 1
                        break

                    list_of_keys[current_table].append(col_value)

                    current_col += 1

            if coordinates[current_table]['start_row'] is None:
                coordinates[current_table]['start_row'] = current_row + 1
                coordinates[current_table]['start_column'] = 1
            coordinates[current_table]['end_row'] = current_row
            coordinates[current_table]['end_column'] = current_col
            current_row += 1

        # Update the `info_to_update` with the info from the file
        for table in list_of_keys.keys():

            for row in range(coordinates[table]['start_row'], coordinates[table]['end_row'] + 1):
                # 1 is first row, not 0
                info_to_update[row] = {key: None for key in list_of_keys[table]}

                for col in range(coordinates[table]['start_column'], coordinates[table]['end_column'] + 1):
                    # 1 is first col, not 0
                    char = get_column_letter(col)
                    # == chr(65 + col)
                    info_to_update[row][list_of_keys[table][col-1]] = worksheet[char + str(row)].value

        # Update all table 1 by 1

            if table == 'AccessLevels':
                AccessLevels.objects.bulk_create([AccessLevels(
                    code=info_to_update[obj]['code'],
                    description=info_to_update[obj]['description'],
                ) for obj in info_to_update.keys()])

            elif table == 'AccessRights':
                AccessRights.objects.bulk_create([AccessRights(
                    type=info_to_update[obj]['type'],
                    description=info_to_update[obj]['description'],
                ) for obj in info_to_update.keys()])

            elif table == 'Trainings':
                Trainings.objects.bulk_create([Trainings(
                    code=info_to_update[obj]['code'],
                    name=info_to_update[obj]['name'],
                    description=info_to_update[obj]['description'],
                    # slug=slugify(f"{info_to_update[obj]['code']}"),
                    slug=slugify(f"{info_to_update[obj]['code']}-"
                                 f"{translate_to_maimunica(info_to_update[obj]['name'])}"),
                ) for obj in info_to_update.keys()])

            elif table == 'Positions':
                Positions.objects.bulk_create([Positions(
                    code=info_to_update[obj]['code'],
                    name=info_to_update[obj]['name'],
                    access_rights=AccessRights.objects.all()[0],  # ToDo: hardcoded for the excel upload
                ) for obj in info_to_update.keys()])

            elif table == 'Employee':
                Employee.objects.bulk_create([Employee(
                    first_name=info_to_update[row_number]['first_name'],
                    middle_name=info_to_update[row_number]['middle_name'],
                    last_name=info_to_update[row_number]['last_name'],
                    identification=info_to_update[row_number]['identification'],

                    position=Positions.objects.all()[0],  # ToDo: hardcoded for the excel upload

                    contract_number=info_to_update[row_number]['contract_number'],
                    starting_date=info_to_update[row_number]['starting_date'],
                    date_last_hse_training=info_to_update[row_number]['date_last_hse_training'],
                    date_next_hse_training=info_to_update[row_number]['date_next_hse_training'],
                    egn=info_to_update[row_number]['egn'],
                    # slug=slugify(f"{info_to_update[row_number]['first_name']}-{info_to_update[row_number]['last_name']}"),
                    # slug=slugify(f"{info_to_update[row_number]['identification']}-"
                    #              f"{info_to_update[row_number]['position']}"),
                    slug=slugify(f"{info_to_update[row_number]['identification']}-"
                                 f"{translate_to_maimunica(info_to_update[row_number]['first_name'])}-"
                                 f"{translate_to_maimunica(info_to_update[row_number]['last_name'])}"),
                )
                    for row_number in info_to_update.keys()])

            elif table == 'Document':
                Document.objects.bulk_create([Document(
                    type=info_to_update[obj]['type'],
                    number=info_to_update[obj]['number'],
                    name=info_to_update[obj]['name'],
                    revision=info_to_update[obj]['revision'],
                    creation_date=info_to_update[obj]['creation_date'],
                    revision_date=info_to_update[obj]['revision_date'],
                    revision_details=info_to_update[obj]['revision_details'],
                    status=info_to_update[obj]['status'],
                    owner=Employee.objects.all()[0],  # ToDo: hardcoded for the excel upload
                    # slug=slugify(f"{info_to_update[obj]['name']}"),
                    # slug=slugify(f"{info_to_update[obj]['owner']}-"
                    #              f"{info_to_update[obj]['type']}-"
                    #              f"{info_to_update[obj]['revision']}"),
                    slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'])}"),
                ) for obj in info_to_update.keys()])

            elif table == 'Process':
                Process.objects.bulk_create([Process(
                    type=info_to_update[obj]['type'],
                    number=info_to_update[obj]['number'],
                    name=info_to_update[obj]['name'],
                    process_owner=Employee.objects.all()[1],  # ToDo: hardcoded for the excel upload 2
                    # slug=slugify(f"{info_to_update[obj]['number']}-"
                    #              f"{info_to_update[obj]['type']}"),
                    slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'])}"),
                ) for obj in info_to_update.keys()])

            elif table == 'ProcessStep':
                ProcessStep.objects.bulk_create([ProcessStep(
                    type=info_to_update[obj]['type'],
                    number=info_to_update[obj]['number'],
                    name=info_to_update[obj]['name'],
                    # parent_process=Process.objects.all()[0],    # ToDo: hardcoded for the excel upload

                    # use the `PXX` from the excel
                    # parent_process=Process.objects.all()[int(info_to_update[obj]['parent_process'][-1])-1],
                    parent_process=Process.objects.all()[int(info_to_update[obj]['parent_process'])-1],

                    description=info_to_update[obj]['description'],
                    responsible=Employee.objects.all()[1],  # ToDo: hardcoded for the excel upload 2
                    # slug=slugify(f"{info_to_update[obj]['name']}"),
                    # slug=slugify(f"{info_to_update[obj]['parent_process']}-"
                    #              f"{info_to_update[obj]['number']}"),
                    slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:50])}"),
                ) for obj in info_to_update.keys()])

            elif table == 'Requirements':
                Requirements.objects.bulk_create([Requirements(
                    organization=info_to_update[obj]['organization'],
                    external_document=info_to_update[obj]['external_document'],
                    clause=info_to_update[obj]['clause'],
                    description=info_to_update[obj]['description'],
                    slug=slugify(f"{info_to_update[obj]['organization']}-"
                                 f"{info_to_update[obj]['clause']}-"
                                 f"{translate_to_maimunica(info_to_update[obj]['description'][0:20])}"),
                ) for obj in info_to_update.keys()])

            info_to_update = {}

        return 'Successfully added \n' \
               '(*documents have no attachments and \n' \
               'process steps have no linked documents!)'

    # Decorator + function: time measurement and writing to a log
    # -----------------------------------------------------------------------

    @staticmethod
    def log_entry(command):
        def decorator_depending_on_command(some_function):
            @wraps(some_function)
            def wrapper(*args, **kwargs):
                if command:
                    start = time.time()
                    result = some_function(*args, **kwargs)  # the function
                    end = time.time()
                    measurement = end - start

                    if custom_collections.logging_info_stack:
                        CaptainsLog.objects.create(
                            operation=f"{custom_collections.logging_info_stack.pop()} "
                                      f"with `{some_function.__name__}`",
                            # performed_at_time=end, # auto added in model
                            execution_time=f"{measurement:.5f} s",
                        )
                else:
                    result = some_function(*args, **kwargs)  # the function
                return result  # the function
            return wrapper
        return decorator_depending_on_command

    @staticmethod
    def log_info(function_result):
        custom_collections.logging_info_stack.append(function_result)

    # new document revision
    # -----------------------------------------------------------------------
    @staticmethod
    def new_revision(the_form):
        result = DocumentEditPurgatory.objects.create(
            type=the_form.cleaned_data['type'],
            number=the_form.cleaned_data['number'],
            name=the_form.cleaned_data['name'],
            revision=the_form.cleaned_data['revision'],
            creation_date=the_form.cleaned_data['creation_date'],
            revision_date=the_form.cleaned_data['revision_date'],
            revision_details=the_form.cleaned_data['revision_details'],
            status=the_form.cleaned_data['status'],
            owner=the_form.cleaned_data['owner'],
            # slug=slugify(f"{info_to_update[obj]['name']}"),
            slug=slugify(f"{the_form.cleaned_data['owner']}-"
                         f"{the_form.cleaned_data['type']}-"
                         f"{the_form.cleaned_data['revision']}"),
        )
        return result

    # search results on the index page
    # -----------------------------------------------------------------------
    @staticmethod
    def search_results(search_pattern, choice):
        result = {}

        if choice == 'All':
            processes = Process.objects.filter(name__icontains=search_pattern)
            process_steps = ProcessStep.objects.filter(name__icontains=search_pattern)
            documents = Document.objects.filter(name__icontains=search_pattern)

            employees_by_first_names = list(Employee.objects.filter(first_name__icontains=search_pattern))
            employees_by_middle_names = list(Employee.objects.filter(middle_name__icontains=search_pattern))
            employees_by_last_names = list(Employee.objects.filter(last_name__icontains=search_pattern))
            employees_temp_list = employees_by_first_names + employees_by_middle_names + employees_by_last_names
            employee_set = set(employees_temp_list)

            trainings = Trainings.objects.filter(name__icontains=search_pattern)

            result['processes'] = processes
            result['process_steps'] = process_steps
            result['documents'] = documents
            result['employees'] = employee_set
            result['trainings'] = trainings

        elif choice == 'Process':
            processes = Process.objects.filter(name__icontains=search_pattern)
            result['processes'] = processes

        elif choice == 'ProcessStep':
            process_steps = ProcessStep.objects.filter(name__icontains=search_pattern)
            result['process_steps'] = process_steps

        elif choice == 'Document':
            documents = Document.objects.filter(name__icontains=search_pattern)
            result['documents'] = documents

        elif choice == 'Employee':
            employees_by_first_names = list(Employee.objects.filter(first_name__icontains=search_pattern))
            employees_by_middle_names = list(Employee.objects.filter(middle_name__icontains=search_pattern))
            employees_by_last_names = list(Employee.objects.filter(last_name__icontains=search_pattern))
            employees_temp_list = employees_by_first_names + employees_by_middle_names + employees_by_last_names
            employee_set = set(employees_temp_list)

            result['employees'] = employee_set

        elif choice == 'Trainings':
            trainings = Trainings.objects.filter(name__icontains=search_pattern)
            result['trainings'] = trainings

        return result
