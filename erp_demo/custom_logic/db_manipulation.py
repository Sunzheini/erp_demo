from django.utils.text import slugify
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from erp_demo.actionplan_mng.models import ActionPlan, ActionPlanStep
from erp_demo.calibration_mng.models import MeasuringEquipment
from erp_demo.characteristics_mng.models import Characteristic
from erp_demo.control_plan_mng.models import ProcessControlPlan, ProcessControlPlanStep
from erp_demo.customer_mng.models import Customer
from erp_demo.defect_cat_mng.models import DefectCatalogue
from erp_demo.dox_mng.models import Document, DocumentLikesToUsers
from erp_demo.hr_mng.models import Employee, Positions, AccessLevels, \
    AccessRights, PositionsToAccessLevels, Trainings, EmployeeToTrainings
from erp_demo.custom_logic import custom_collections
from erp_demo.interaction_mng.models import InteractionToDocuments, Interaction
from erp_demo.kpi_mng.models import Kpi
from erp_demo.main_app.models import CaptainsLog, Requirements
from erp_demo.custom_logic.translator import translate_to_maimunica
from erp_demo.maintenance_mng.models import Machine
from erp_demo.newactions_mng.models import NewAction
from erp_demo.nonconformity_mng.models import Nonconformity
from erp_demo.opportunity_mng.models import Opportunity
from erp_demo.organization_mng.models import Organization
from erp_demo.process_mng.models import ProcessStepToDocuments, \
    ProcessStep, Process, ProcessToKpis, ProcessToOpportunities, ProcessToRisks
from erp_demo.resource_mng.models import Resource, ResourcesAssignedToEmployees, ResourcesAssignedToProcess
from erp_demo.review_mng.models import ManagementReview
from erp_demo.risk_mng.models import Risk
from erp_demo.statistics_mng.models import StatModel1
from erp_demo.supplier_mng.models import Supplier, Material


class DatabaseManipulation:

# Delete whole db (the order is chosen in order not have issues with the links between tables)
# -----------------------------------------------------------------------

    @staticmethod
    def delete_database():
        # M2M tables
        ProcessStepToDocuments.objects.all().delete()
        PositionsToAccessLevels.objects.all().delete()
        EmployeeToTrainings.objects.all().delete()
        DocumentLikesToUsers.objects.all().delete()
        ProcessToKpis.objects.all().delete()
        ProcessToOpportunities.objects.all().delete()
        ProcessToRisks.objects.all().delete()
        InteractionToDocuments.objects.all().delete()
        ResourcesAssignedToEmployees.objects.all().delete()
        ResourcesAssignedToProcess.objects.all().delete()

        # tables with no dependencies to other tables
        AccessLevels.objects.all().delete()
        AccessRights.objects.all().delete()
        Trainings.objects.all().delete()
        Interaction.objects.all().delete()
        Resource.objects.all().delete()
        NewAction.objects.all().delete()
        MeasuringEquipment.objects.all().delete()
        Machine.objects.all().delete()
        Characteristic.objects.all().delete()
        StatModel1.objects.all().delete()
        Material.objects.all().delete()

        # tables with dependencies from other tables
        Positions.objects.all().delete()
        Employee.objects.all().delete()
        ProcessStep.objects.all().delete()
        Document.objects.all().delete()
        Kpi.objects.all().delete()
        Opportunity.objects.all().delete()
        Risk.objects.all().delete()
        Process.objects.all().delete()
        CaptainsLog.objects.all().delete()
        Requirements.objects.all().delete()
        Organization.objects.all().delete()
        Customer.objects.all().delete()
        Nonconformity.objects.all().delete()
        ActionPlanStep.objects.all().delete()
        ActionPlan.objects.all().delete()
        Supplier.objects.all().delete()
        ProcessControlPlanStep.objects.all().delete()
        ProcessControlPlan.objects.all().delete()
        DefectCatalogue.objects.all().delete()
        ManagementReview.objects.all().delete()

        return 'Successfully deleted'


# Upload to db
# -----------------------------------------------------------------------

    @staticmethod
    def recreate_database(request_object):
        # Select the file
        current_file_path = request_object['select_file']
        workbook = load_workbook(current_file_path)
        worksheet = workbook[workbook.sheetnames[0]]

        # Get the collections
        info_to_update = {}
        list_of_keys = custom_collections.list_of_keys
        coordinates = custom_collections.coordinates

        # Fill the coordinates
        consecutive_empty_rows = 0
        current_row = 1
        current_table = ''
        is_table_header_done = False
        while consecutive_empty_rows < 2:
            row_value = worksheet['A' + str(current_row)].value

            if row_value is None:
                consecutive_empty_rows += 1
                current_row += 1
                is_table_header_done = False
                current_table = ''
                continue

            if current_table == '':
                if row_value in coordinates.keys():
                    current_table = row_value
                    consecutive_empty_rows = 0
                    current_row += 1
                    continue

            if not is_table_header_done:
                current_col = 1
                while 1:
                    col_value = worksheet[get_column_letter(current_col) + str(current_row)].value
                    if col_value is None:
                        is_table_header_done = True
                        current_col -= 1
                        break

                    list_of_keys[current_table].append(col_value)

                    current_col += 1

            if coordinates[current_table]['start_row'] is None:
                coordinates[current_table]['start_row'] = current_row + 1
                coordinates[current_table]['start_column'] = 1
            coordinates[current_table]['end_row'] = current_row
            coordinates[current_table]['end_column'] = current_col
            current_row += 1

        # Update the `info_to_update` with the info from the file
        for table in list_of_keys.keys():

            for row in range(coordinates[table]['start_row'], coordinates[table]['end_row'] + 1):
                # 1 is first row, not 0
                info_to_update[row] = {key: None for key in list_of_keys[table]}

                for col in range(coordinates[table]['start_column'], coordinates[table]['end_column'] + 1):
                    # 1 is first col, not 0
                    char = get_column_letter(col)
                    # == chr(65 + col)
                    info_to_update[row][list_of_keys[table][col-1]] = worksheet[char + str(row)].value

        # Update all tables 1 by 1
            if table == 'AccessLevels':
                try:
                    AccessLevels.objects.bulk_create([AccessLevels(
                        code=info_to_update[obj]['code'],
                        description=info_to_update[obj]['description'],
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'AccessRights':
                try:
                    AccessRights.objects.bulk_create([AccessRights(
                        type=info_to_update[obj]['type'],
                        description=info_to_update[obj]['description'],
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Trainings':
                try:
                    Trainings.objects.bulk_create([Trainings(
                        code=info_to_update[obj]['code'],
                        name=info_to_update[obj]['name'],
                        description=info_to_update[obj]['description'],
                        # slug=slugify(f"{info_to_update[obj]['code']}"),
                        slug=slugify(f"{info_to_update[obj]['code']}-"
                                     f"{translate_to_maimunica(info_to_update[obj]['name'])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Positions':
                try:
                    Positions.objects.bulk_create([Positions(
                        code=info_to_update[obj]['code'],
                        name=info_to_update[obj]['name'],
                        access_rights=AccessRights.objects.all()[0],  # ToDo: hardcoded for the excel upload
                        slug=slugify(f"{info_to_update[obj]['code']}-{info_to_update[obj]['name']}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Employee':
                try:
                    Employee.objects.bulk_create([Employee(
                        first_name=info_to_update[row_number]['first_name'],
                        middle_name=info_to_update[row_number]['middle_name'],
                        last_name=info_to_update[row_number]['last_name'],
                        identification=info_to_update[row_number]['identification'],

                        position=Positions.objects.all()[0],  # ToDo: hardcoded for the excel upload

                        contract_number=info_to_update[row_number]['contract_number'],
                        starting_date=info_to_update[row_number]['starting_date'],
                        date_last_hse_training=info_to_update[row_number]['date_last_hse_training'],
                        date_next_hse_training=info_to_update[row_number]['date_next_hse_training'],
                        egn=info_to_update[row_number]['egn'],
                        slug=slugify(f"{info_to_update[row_number]['identification']}-"
                                     f"{translate_to_maimunica(info_to_update[row_number]['first_name'])}-"
                                     f"{translate_to_maimunica(info_to_update[row_number]['last_name'])}"),
                    )
                        for row_number in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Document':
                try:
                    Document.objects.bulk_create([Document(
                        type=info_to_update[obj]['type'],
                        number=info_to_update[obj]['number'],
                        name=info_to_update[obj]['name'],
                        revision=info_to_update[obj]['revision'],
                        creation_date=info_to_update[obj]['creation_date'],
                        revision_date=info_to_update[obj]['revision_date'],
                        revision_details=info_to_update[obj]['revision_details'],
                        status=info_to_update[obj]['status'],
                        owner=Employee.objects.all()[0],  # ToDo: hardcoded for the excel upload
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:50])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Process':
                try:
                    Process.objects.bulk_create([Process(
                        type=info_to_update[obj]['type'],
                        number=info_to_update[obj]['number'],
                        name=info_to_update[obj]['name'],
                        process_owner=Employee.objects.all()[1],  # ToDo: hardcoded for the excel upload 2
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:50])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'ProcessStep':
                try:
                    ProcessStep.objects.bulk_create([ProcessStep(
                        type=info_to_update[obj]['type'],
                        number=info_to_update[obj]['number'],
                        name=info_to_update[obj]['name'],
                        # parent_process=Process.objects.all()[0],    # ToDo: hardcoded for the excel upload
                        parent_process=Process.objects.all()[int(info_to_update[obj]['parent_process']) - 1],

                        description=info_to_update[obj]['description'],
                        responsible=Employee.objects.all()[1],  # ToDo: hardcoded for the excel upload 2
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:50])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Requirements':
                try:
                    Requirements.objects.bulk_create([Requirements(
                        organization=info_to_update[obj]['organization'],
                        external_document=info_to_update[obj]['external_document'],
                        clause=info_to_update[obj]['clause'],
                        clause_name=info_to_update[obj]['clause_name'],
                        description=info_to_update[obj]['description'],
                        slug=slugify(f"{info_to_update[obj]['organization']}-"
                                     f"{info_to_update[obj]['clause']}-"
                                     f"{translate_to_maimunica(info_to_update[obj]['description'][0:20])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Organization':
                try:
                    Organization.objects.bulk_create([Organization(
                        name=info_to_update[obj]['name'],
                        eik=info_to_update[obj]['eik'],
                        mol=info_to_update[obj]['mol'],
                        address=info_to_update[obj]['address'],
                        manager_name=info_to_update[obj]['manager_name'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:20])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Customer':
                try:
                    Customer.objects.bulk_create([Customer(
                        type=info_to_update[obj]['type'],
                        name=info_to_update[obj]['name'],
                        registration_address=info_to_update[obj]['registration_address'],
                        registration_city=info_to_update[obj]['registration_city'],
                        eik=info_to_update[obj]['eik'],
                        mol1=info_to_update[obj]['mol1'],
                        mol2=info_to_update[obj]['mol2'],
                        mol3=info_to_update[obj]['mol3'],
                        mol4=info_to_update[obj]['mol4'],
                        mol5=info_to_update[obj]['mol5'],
                        correspondence_address1=info_to_update[obj]['correspondence_address1'],
                        correspondence_address2=info_to_update[obj]['correspondence_address2'],
                        correspondence_address3=info_to_update[obj]['correspondence_address3'],
                        correspondence_address4=info_to_update[obj]['correspondence_address4'],
                        correspondence_address5=info_to_update[obj]['correspondence_address5'],
                        contact_person1=info_to_update[obj]['contact_person1'],
                        contact_person2=info_to_update[obj]['contact_person2'],
                        contact_person3=info_to_update[obj]['contact_person3'],
                        contact_person4=info_to_update[obj]['contact_person4'],
                        contact_person5=info_to_update[obj]['contact_person5'],
                        phone1=info_to_update[obj]['phone1'],
                        phone2=info_to_update[obj]['phone2'],
                        phone3=info_to_update[obj]['phone3'],
                        phone4=info_to_update[obj]['phone4'],
                        phone5=info_to_update[obj]['phone5'],
                        email1=info_to_update[obj]['email1'],
                        email2=info_to_update[obj]['email2'],
                        email3=info_to_update[obj]['email3'],
                        email4=info_to_update[obj]['email4'],
                        email5=info_to_update[obj]['email5'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:20])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Kpi':
                try:
                    Kpi.objects.bulk_create([Kpi(
                        name=info_to_update[obj]['name'],
                        description=info_to_update[obj]['description'],
                        target=info_to_update[obj]['target'],
                        actual_01_23=info_to_update[obj]['actual_01_23'],
                        actual_02_23=info_to_update[obj]['actual_02_23'],
                        actual_03_23=info_to_update[obj]['actual_03_23'],
                        actual_04_23=info_to_update[obj]['actual_04_23'],
                        actual_05_23=info_to_update[obj]['actual_05_23'],
                        actual_06_23=info_to_update[obj]['actual_06_23'],
                        actual_07_23=info_to_update[obj]['actual_07_23'],
                        actual_08_23=info_to_update[obj]['actual_08_23'],
                        actual_09_23=info_to_update[obj]['actual_09_23'],
                        actual_10_23=info_to_update[obj]['actual_10_23'],
                        actual_11_23=info_to_update[obj]['actual_11_23'],
                        actual_12_23=info_to_update[obj]['actual_12_23'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:20])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Opportunity':
                try:
                    Opportunity.objects.bulk_create([Opportunity(
                        name=info_to_update[obj]['name'],
                        description=info_to_update[obj]['description'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:20])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Risk':
                try:
                    Risk.objects.bulk_create([Risk(
                        name=info_to_update[obj]['name'],
                        description=info_to_update[obj]['description'],
                        probability=info_to_update[obj]['probability'],
                        impact=info_to_update[obj]['impact'],
                        immediate_action=info_to_update[obj]['immediate_action'],
                        ia_test_period=info_to_update[obj]['ia_test_period'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:20])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Interaction':
                try:
                    Interaction.objects.bulk_create([Interaction(
                        name=info_to_update[obj]['name'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:20])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Resource':
                try:
                    Resource.objects.bulk_create([Resource(
                        name=info_to_update[obj]['name'],
                        description=info_to_update[obj]['description'],
                        quantity=info_to_update[obj]['quantity'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:20])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Nonconformity':
                try:
                    Nonconformity.objects.bulk_create([Nonconformity(
                        name=info_to_update[obj]['name'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:30])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Supplier':
                try:
                    Supplier.objects.bulk_create([Supplier(
                        name=info_to_update[obj]['name'],
                        description=info_to_update[obj]['description'],
                        score=info_to_update[obj]['score'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:30])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'MeasuringEquipment':
                try:
                    MeasuringEquipment.objects.bulk_create([MeasuringEquipment(
                        name=info_to_update[obj]['name'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:30])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Machine':
                try:
                    Machine.objects.bulk_create([Machine(
                        name=info_to_update[obj]['name'],
                        inventory_number=info_to_update[obj]['inventory_number'],
                        characteristics=info_to_update[obj]['characteristics'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:30])}"),
                    ) for obj in info_to_update.keys()])

                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Characteristic':
                try:
                    Characteristic.objects.bulk_create([Characteristic(
                        name=info_to_update[obj]['name'],
                        code=info_to_update[obj]['code'],
                        type=info_to_update[obj]['type'],
                        requirement=info_to_update[obj]['requirement'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:30])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'ProcessControlPlan':
                try:
                    ProcessControlPlan.objects.bulk_create([ProcessControlPlan(
                        name=info_to_update[obj]['name'],
                        type=info_to_update[obj]['type'],
                        number=info_to_update[obj]['number'],
                        revision=info_to_update[obj]['revision'],
                        product=info_to_update[obj]['product'],
                        team=info_to_update[obj]['team'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:30])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'ProcessControlPlanStep':
                try:
                    ProcessControlPlanStep.objects.bulk_create([ProcessControlPlanStep(
                        name=info_to_update[obj]['name'],
                        sample_size=info_to_update[obj]['sample_size'],
                        frequency=info_to_update[obj]['frequency'],
                        reaction_plan=info_to_update[obj]['reaction_plan'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:30])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'DefectCatalogue':
                try:
                    DefectCatalogue.objects.bulk_create([DefectCatalogue(
                        name=info_to_update[obj]['name'],
                        number=info_to_update[obj]['number'],
                        description=info_to_update[obj]['description'],
                        slug=slugify(f"{translate_to_maimunica(info_to_update[obj]['name'][0:30])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'StatModel1':
                try:
                    StatModel1.objects.bulk_create([StatModel1(
                        name=info_to_update[obj]['name'],
                        operator=info_to_update[obj]['operator'],
                        grinding=info_to_update[obj]['grinding'],
                        welding=info_to_update[obj]['welding'],
                        blasting=info_to_update[obj]['blasting'],
                        painting=info_to_update[obj]['painting'],
                        assembly=info_to_update[obj]['assembly'],
                        total_pieces=info_to_update[obj]['total_pieces'],
                        slug=slugify(f"{translate_to_maimunica(str(info_to_update[obj]['name'])[0:30])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'ManagementReview':
                try:
                    ManagementReview.objects.bulk_create([ManagementReview(
                        name=info_to_update[obj]['name'],
                        date=info_to_update[obj]['date'],
                        slug=slugify(f"{translate_to_maimunica(str(info_to_update[obj]['name'])[0:30])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            elif table == 'Material':
                try:
                    Material.objects.bulk_create([Material(
                        name=info_to_update[obj]['name'],
                        description=info_to_update[obj]['description'],
                        quantity=info_to_update[obj]['quantity'],
                        measurement_unit=info_to_update[obj]['measurement_unit'],
                        price_per_unit=info_to_update[obj]['price_per_unit'],
                        slug=slugify(f"{translate_to_maimunica(str(info_to_update[obj]['name'])[0:30])}"),
                    ) for obj in info_to_update.keys()])
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    pass

            info_to_update = {}

        return 'Successfully added \n' \
               '(*documents have no attachments and \n' \
               'process steps have no linked documents!)'
