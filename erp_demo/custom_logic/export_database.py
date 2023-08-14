from openpyxl import Workbook
from django.db import models

from django.http import HttpResponse
from io import BytesIO
import os
from django.conf import settings
from tempfile import NamedTemporaryFile
from tempfile import TemporaryDirectory
import shutil

from openpyxl.writer.excel import save_virtual_workbook

from erp_demo.hr_mng.models import AccessLevels
from erp_demo.hr_mng.models import AccessRights
from erp_demo.hr_mng.models import Trainings
from erp_demo.hr_mng.models import Positions
from erp_demo.hr_mng.models import Employee
from erp_demo.dox_mng.models import Document
from erp_demo.process_mng.models import Process
from erp_demo.process_mng.models import ProcessStep
from erp_demo.main_app.models import Requirements
from erp_demo.organization_mng.models import Organization
from erp_demo.customer_mng.models import Customer
from erp_demo.kpi_mng.models import Kpi
from erp_demo.opportunity_mng.models import Opportunity
from erp_demo.risk_mng.models import Risk
from erp_demo.interaction_mng.models import Interaction
from erp_demo.resource_mng.models import Resource
from erp_demo.nonconformity_mng.models import Nonconformity
from erp_demo.supplier_mng.models import Supplier
from erp_demo.calibration_mng.models import MeasuringEquipment
from erp_demo.maintenance_mng.models import Machine
from erp_demo.characteristics_mng.models import Characteristic
from erp_demo.control_plan_mng.models import ProcessControlPlan
from erp_demo.control_plan_mng.models import ProcessControlPlanStep
from erp_demo.defect_cat_mng.models import DefectCatalogue
from erp_demo.statistics_mng.models import StatModel1
from erp_demo.review_mng.models import ManagementReview


LIST_OF_MODELS_TO_EXTRACT = [
    AccessLevels,
    AccessRights,
    Trainings,
	Positions,
	Employee,
    Document,
    Process,
    ProcessStep,
    Requirements,
    Organization,
    Customer,
    Kpi,
    Opportunity,
    Risk,
    Interaction,
    Resource,
    Nonconformity,
    Supplier,
    MeasuringEquipment,
    Machine,
    Characteristic,
    ProcessControlPlan,
    ProcessControlPlanStep,
    DefectCatalogue,
    StatModel1,
    ManagementReview,
]


# def correct_data_in_the_row(row):
#     for data in row:
#         if data == 'None' or 'None' in data:
#             row[row.index(data)] = ''
#
#     return row
#
#
# def export_database():
#     workbook = Workbook()
#
#     for model in LIST_OF_MODELS_TO_EXTRACT:
#         model_name = model.__name__
#         worksheet = workbook.create_sheet(model_name)
#
#         objects = model.objects.all()
#
#         model_instance = model()
#         headers = [field.name for field in model_instance._meta.get_fields()
#                    if isinstance(field, models.Field) and field.name != 'slug']
#
#         worksheet.append(headers)
#
#         for obj in objects:
#             data_row = [str(getattr(obj, field_name)) for field_name in headers]
#             corrected_data = correct_data_in_the_row(data_row)
#             worksheet.append(corrected_data)
#
#     workbook.remove(workbook.active)  # Remove the default sheet created by openpyxl
#     file_path = r"C:\Users\User\Desktop\exported_data.xlsx"
#     workbook.save(file_path)
#
#     return f'Successfully exported to excel file to {file_path}'


def correct_data_in_the_row(row):
    for data in row:
        if data == 'None' or 'None' in data:
            row[row.index(data)] = ''

    return row


def ensure_directory_exists(directory_path):
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)


def export_database():
    workbook = Workbook()

    for model in LIST_OF_MODELS_TO_EXTRACT:
        model_name = model.__name__
        worksheet = workbook.create_sheet(model_name)

        objects = model.objects.all()

        model_instance = model()
        headers = [field.name for field in model_instance._meta.get_fields()
                   if isinstance(field, models.Field) and field.name != 'slug']

        worksheet.append(headers)

        for obj in objects:
            data_row = [str(getattr(obj, field_name)) for field_name in headers]
            corrected_data = correct_data_in_the_row(data_row)
            worksheet.append(corrected_data)

    workbook.remove(workbook.active)  # Remove the default sheet created by openpyxl

    # Create a temporary file
    with NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        workbook.save(temp_file)

    # Close the temporary file before moving it
    temp_file.close()

    # Ensure the target directory exists
    target_directory = os.path.join(settings.STATIC_ROOT, 'excel')
    ensure_directory_exists(target_directory)

    # Move the temporary file to the target location
    final_file_path = os.path.join(target_directory, 'exported_data.xlsx')
    shutil.move(temp_file.name, final_file_path)

    return f'Successfully exported to excel file to {final_file_path}'
