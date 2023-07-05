from openpyxl import load_workbook
from django.apps import apps


def extract_from_excel(request_object, app_name, model_name, rows_dict):
    # Select the file
    current_file_path = request_object['select_file']
    workbook = load_workbook(current_file_path)
    worksheet = workbook[workbook.sheetnames[1]]
    new_dict = {}

    for key, value in rows_dict.items():
        row_value = worksheet[value].value
        if row_value:
            new_dict[key] = str(row_value)

    model_class = apps.get_model(app_name, model_name)
    model_class.objects.create( **new_dict)
