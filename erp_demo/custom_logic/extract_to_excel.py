from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from django.db.models.fields.related import ForeignKey, ManyToManyField
from django.db.models import Field


class ExtractToExcel:
    def __init__(self, file_path, object_to_extract, additional_object_list=None, fields=None):
        self.file_path = file_path
        self.object_to_extract = object_to_extract
        self.all_fields = [f for f in self.object_to_extract._meta.get_fields() if isinstance(f, Field)]
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.additional = additional_object_list
        self.fields = fields

    @staticmethod
    def _get_column_letter(column_number):
        return get_column_letter(column_number)

    def write_to_cell(self, row, column, value):
        self.worksheet.cell(row=row, column=column, value=value)
        self.workbook.save(self.file_path)

    def run(self):
        current_row = 1
        for field in self.all_fields:
            current_row += 1
            value = getattr(self.object_to_extract, field.name)

            # Handle foreign key and many to many fields
            if isinstance(field, ForeignKey):
                value = value.pk if value else None
            elif isinstance(field, ManyToManyField):
                value = ', '.join(str(item.pk) for item in value.all()) if value else None

            self.write_to_cell(current_row, 1, field.name)
            self.write_to_cell(current_row, 2, str(value) if value else None)

        if self.additional:
            for obj in self.additional:
                current_row += 1
                for field in self.fields:
                    self.write_to_cell(current_row, self.fields.index(field) + 1, getattr(obj, field))
